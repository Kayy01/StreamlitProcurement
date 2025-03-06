import os
import streamlit as st
import pdfplumber
import docx
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from dotenv import load_dotenv
from langchain_openai import AzureChatOpenAI
from langchain.schema import HumanMessage
from langchain.text_splitter import RecursiveCharacterTextSplitter

# Load environment variables
load_dotenv(dotenv_path="index.env", override=True)
OPENAI_DEPLOYMENT_NAME = os.getenv("OPENAI_DEPLOYMENT_NAME")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

if not all([OPENAI_DEPLOYMENT_NAME, AZURE_OPENAI_ENDPOINT, OPENAI_API_KEY]):
    st.error("Missing required environment variables. Check your .env file.")
    st.stop()

# Initialize Azure OpenAI Chat model
llm = AzureChatOpenAI(
    azure_deployment=OPENAI_DEPLOYMENT_NAME,
    azure_endpoint=f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{OPENAI_DEPLOYMENT_NAME}/chat/completions?api-version=2024-10-21",
    openai_api_key=OPENAI_API_KEY,
    openai_api_version="2024-10-21"
)

# Function to extract text from PDF
def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text.strip()

# Function to extract text from DOCX
def extract_text_from_docx(docx_path):
    doc = docx.Document(docx_path)
    return "\n".join([p.text.strip() for p in doc.paragraphs if p.text.strip()])

# Function to process and extract structured information
def process_document(file):
    file_extension = file.name.split(".")[-1]
    text = extract_text_from_pdf(file) if file_extension == "pdf" else extract_text_from_docx(file) if file_extension == "docx" else ""

    if not text:
        return None
    
    # Splitting text for processing
    splitter = RecursiveCharacterTextSplitter(chunk_size=5000, chunk_overlap=0)
    text_chunks = splitter.split_text(text)
    extracted_data = []
    
    for chunk in text_chunks:
        prompt = f"""
        Analyze the document properly, extract main & subitems properly and their descriptions from the following text:
        {chunk}
        
        **Formatting Rules:**
        - Extract all item names and descriptions in a structured manner. DO NOT add any extra comments!
        - Do not extract Price!
        - Maintain logical sentence breaks.
        - Format the extracted data properly for user readability.
        - Each line should have max **45 characters**.
        - Do **not** split words in half when breaking lines.
        - If exceeding 45 characters, split into a new row.

        Do not add any extra symbols before the answer.
        Do not add unnecessary gaps in between output.
        Do not add any extra words/comment in the answer. Strictly follow formatting rules.
        """
        
        response = llm.invoke([HumanMessage(content=prompt)])
        extracted_data.append(response.content.strip())
    
    return extracted_data

# Streamlit UI
st.title("📄 AI-Driven Information Extractor")

uploaded_file = st.file_uploader("Upload a PDF or DOCX file", type=["pdf", "docx"])
excel_template = st.file_uploader("Upload Excel Template", type=["xlsx"])

# Store extracted data in session state to prevent re-extraction
if uploaded_file:
    if "extracted_data" not in st.session_state or st.session_state["file_name"] != uploaded_file.name:
        st.session_state["extracted_data"] = process_document(uploaded_file)
        st.session_state["file_name"] = uploaded_file.name  # Track file change

# Display extracted data & allow editing
if "extracted_data" in st.session_state and st.session_state["extracted_data"]:
    st.subheader("📋 Extracted Data Review")

    # Convert list to a single editable string
    formatted_text = "\n".join(st.session_state["extracted_data"])
    
    # Use st.text_area for user to edit extracted text
    edited_text = st.text_area("Review and Edit the Extracted Data:", formatted_text, height=300)

    # Save edited text back to session state
    st.session_state["edited_data"] = edited_text.split("\n")

    # Fill Excel Template Button
    if excel_template and st.button("Fill Template"):
        wb = load_workbook(excel_template)
        sheet = wb.active

        # Find column indices for 'Description' and 'Description 2'
        col_indices = {col[0].value: col[0].column for col in sheet.iter_cols(1, sheet.max_column)}
        desc_col = col_indices.get("Description")
        desc2_col = col_indices.get("Description 2")

        if desc_col:
            row_idx = 2  # Start from row 2 (assuming row 1 has headers)

            # Use edited data instead of original extracted data
            for row_text in st.session_state["edited_data"]:
                sheet.cell(row=row_idx, column=desc_col, value=row_text).alignment = Alignment(wrap_text=True)
                if desc2_col:
                    sheet.cell(row=row_idx, column=desc2_col, value="")  # Keep Description 2 blank
                row_idx += 1  # Move to the next row

            output_path = "updated_template.xlsx"
            wb.save(output_path)
            st.success("✅ Data filled successfully!")
            st.download_button("📥 Download Updated Excel", data=open(output_path, "rb"), file_name="updated_template.xlsx")
        else:
            st.error("⚠️ 'Description' column not found in Excel template.")
